#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Mar 10 23:53:48 2018

@author: chandrasendr
"""

#! python3
# sendMarketingEmails.py - Sends emails to all email ids in spreadsheet.

import openpyxl, smtplib, sys

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('marketingEmail.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# Get all email id's
mailingList = {}
for r in range(2, sheet.max_row + 1):
    sent = sheet.cell(row=r, column=lastCol).value
    if sent != 'notsent':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        mailingList[name] = email

# Log in to email account.
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
# Enter desired email address as password
smtpObj.login('chandratestacc17@gmail.com', 'testacc@123')
# smtpObj.login(str(input('Please enter the email address: ')), str(input('Please enter the password:')))

# Send out reminder emails.
for name, email in mailingList.items():
    body = """Subject: %s & Nico Ahlers - Vorstellung.\n
Dear %s,\n
    
Während meiner Recherche zu Targecy bin ich auf Ihr Profil gestoßen und denke, Sie sind der richtige Ansprechpartner für die Themen Planung und Organisation der Bürosauberkeit. Daher haben Sie sicherlich bereits ein sauberes Büro, weil Ihnen bewusst ist, wie wichtig eine saubere Umgebung für das Wohlbefinden und damit die Effizienz der Mitarbeiter ist.

Wir von BOOK A TIGER Business sind nicht das am schnellsten wachsende Facility Services Unternehmen, weil wir einfach nur gut sauber machen, sondern weil wir darüber hinaus weitere Probleme von Unternehmen erkannt haben. Oft tauchen neben dem Thema Sauberkeit weitere Fragen auf, die aus Intransparenz und mangelnder Kommunikation resultieren: Wer übernimmt die Urlaubsvertretung unserer Reinigungskraft? Wie lang ist unsere Reinigungskraft pro Termin in unseren Räumen? Wie kann ich einen Termin für das Fensterputzen hinzuzubuchen? 

Wir beantworten diese Fragen ganz einfach! Denn neben Sauberkeit bieten wir mehr Transparenz, bessere Kommunikationsstrukturen und somit mehr Zeitersparnis als andere Unternehmen. Auf unserer Plattform können Sie Ihre professionelle Reinigungsdienstleistung mit ein paar Klicks Planen, Buchen und Kontrollieren. Ebenso einfach können Sie mit Ihrer Reinigungskraft direkt Absprachen treffen. Dank Transparenz und einfachen Kommunikationsmöglichkeiten sparen Sie wertvolle Zeit und können sich wieder auf wichtige Aufgaben fokussieren.

Wären Sie bereit für ein 5 Minuten und 47 Sekunden dauerndes Telefonat diese / nächste Woche, um zu schauen, ob es zwischen Ihren Bedürfnissen und unseren Lösungen Übereinstimmungen gibt?

Viele Grüße,
Nico Ahlers

Nico Ahlers
Senior Sales Manager 
BOOK A TIGER Business

Phone: 49 (0)30 30 80 72 63
Mail: sales@b2b.bookatiger.com

www.bookatiger.com/de-de/business""" %(name, name)
    
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('chandratest17@gmail.com', email, body.encode('utf-8'))
    
    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
        smtpObj.quit()